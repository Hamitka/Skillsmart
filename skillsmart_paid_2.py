"""3. Задания
3.1. Напишите функцию, которая получает на вход имя файла -
документа Word, и две строки, и выполняет в документе замену одной строки на другую.
"""

import os
import random
import pandas as pd
from docx import Document
from openpyxl import Workbook, load_workbook


def doc_replace_string(filename: str, str_find: str, str_replace: str):
    """ Function to replace one string to another in some *.docx file """

    if os.path.isfile:
        document = Document(filename)
    else:
        return False
    for p in document.paragraphs:
        if p.text.find(str_find) != -1:
            p.text = p.text.replace(str_find, str_replace)
            # print(len(p.text), p.text)
    document.save(filename)
    return True

# doc_replace_string('some_files\\test.docx', 'программистом', 'хорошим программистом')
# doc_replace_string('some_files\\test.docx', 'хорошим', 'профессиональным')


""" 3.2. Заполните в Excel-файле на любом листе квадрат 12*12 ячеек случайными значениями. """

filename = 'some_files\\test.xlsx'
wb = load_workbook(filename)
ws = wb['Лист1']
for i in range(1, 13):
    for j in range(1, 13):
        ws.cell(i, j).value = random.randint(0, 1000)
wb.save(filename)

# но по мне так лучше использовать pandas для таких целей:
filename1 = 'some_files\\test1.xlsx'
lst_out = [[random.randint(0, 1000) for i in range(12)] for j in range(12)]
df = pd.DataFrame(lst_out)
df.to_excel(filename1, header=False, index=False)